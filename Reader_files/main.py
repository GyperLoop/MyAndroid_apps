import os
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook

from kivy.app import App
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.popup import Popup
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.lang import Builder
from kivy.uix.textinput import TextInput

Builder.load_file("document_reader.kv")

def read_pdf(path: str) -> str:
    text = []
    with open(path, 'rb') as f:
        reader = PdfReader(f)
        for page in reader.pages:
            text.append(page.extract_text() or "")
    return "\n".join(text)

def read_docx(path: str) -> str:
    doc = Document(path)
    return "\n".join([para.text for para in doc.paragraphs])

def read_txt(path: str) -> str:
    with open(path, 'r', encoding='utf-8') as f:
        return f.read()
    
def read_xlsx(path: str) -> str:
    wb = load_workbook(path, data_only=True)
    output = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        output.append(f"--- Sheet: {sheet_name} ---")
        for row in sheet.iter_rows(value_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row] 
            output.append("\t".join(row_values))
        output.append("")
    return "\n".join(output)

def read_document(path: str) -> str:
    if not os.path.exists(path):
        raise FileNotFoundError("File not found.")
    if os.path.getsize(path) > 50 * 1024 * 1024:
        raise ValueError("File too large.")
    ext = os.path.splitext(path)[1].lower()
    safe_exts = {'.pdf': read_pdf, '.docx': read_docx, '.txt': read_txt, '.xlsx': read_xlsx}
    if ext not in safe_exts:
        raise ValueError(f"Unsupported or unsafe file type: {ext}")
    return safe_exts[ext](path)
    
class DocumentReaderApp(App):
    def build(self):
        return Builder.load_file("document_reader.kv")
    
    def open_file_chooser(self):
        layout = BoxLayout(orientation='vertical', spacing=5, padding=5)
        search_bar = TextInput(
            hint_text="Search by file name...",
            size_hint_y=None,
            height='40dp'
        )
        chooser = FileChooserListView(filters=['*.pdf', '*.docx', '*.txt', '*.xlsx'])
        # chooser.path = "/storage/emulated/0/Documents"  # Optional default path
        btn = Button(
            text='Open Selected',
            size_hint_y=None,
            height='48dp'
        )

        def update_search(instance, value):
            text = value.lower().strip()
            if text:
                chooser.filters = [
                    lambda folder, filename: text in filename.lower() and (
                        filename.endswith('.pdf') or
                        filename.endswith('.docx') or
                        filename.endswith('.txt') or
                        filename.endswith('.xlsx')
                    )
                ]
            else:
                # Reset to show all supported file types
                chooser.filters = ['*.pdf', '*.docx', '*.txt', '*.xlsx']
            chooser._update_files()  # refresh file list

        search_bar.bind(text=update_search)
    
        def open_selected(instance):
            if chooser.selection:
                file_path = chooser.selection[0]
                try:
                    content = read_document(file_path)
                    self.root.ids.output.text = content
                except Exception as e:
                    self.root.ids.output.text = f"Error: {e}"
                popup.dismiss()

        btn.bind(on_release=open_selected)
        layout.add_widget(search_bar)
        layout.add_widget(chooser)
        layout.add_widget(btn)

        popup = Popup(
            title="Select a Document",
            content=layout,
            size_hint=(0.9, 0.9)
        )
        popup.open()     
    
if __name__ == '__main__':
    DocumentReaderApp().run()
    DocumentReaderApp().run()
